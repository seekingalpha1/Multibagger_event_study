"""
Winner Stocks Analysis Script

Analyzes what happens AFTER a stock achieves 2x, 3x, 4x, ... 10x returns.
Answers the question: How long should you hold your winners?

Key Features:
- Uses rolling 5th percentile (252 days) as entry point
- Detects crossing events (from below to above threshold)
- 90-day cooldown period between events
- Tracks performance over 1, 2, 3, 5, and 10 years
- Comprehensive KPIs including returns, drawdowns, probabilities
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
import warnings
from tqdm import tqdm
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import yfinance as yf

warnings.filterwarnings('ignore')


class WinnerStocksAnalyzer:
    """
    Analyzer for winner stocks performance after achieving multiples.
    """
    
    def __init__(self, price_data):
        """
        Initialize analyzer with price data.
        
        Parameters:
        -----------
        price_data : pd.DataFrame
            DataFrame from download_sp500_data.py with columns:
            Date, Ticker, Close, in_sp500, etc.
        """
        self.data = price_data.copy()
        self.data["Date"] = pd.to_datetime(self.data["Date"], utc=True)
        self.data = self.data.sort_values(['Ticker', 'Date']).reset_index(drop=True)
        
        # Configuration
        self.rolling_window = 252  # 1 year for rolling minimum
        self.percentile = 5  # 5th percentile
        self.cooldown_days = 90  # 90 days between events
        self.multiples = [2, 3, 4, 5, 6, 7, 8, 9, 10]  # Multiples to analyze
        self.followup_periods = {
            '1Y': 252,
            '2Y': 504,
            '3Y': 756,
            '5Y': 1260,
            '10Y': 2520
        }
        
        print("Initializing Winner Stocks Analyzer...")
        print(f"Data shape: {self.data.shape}")
        print(f"Date range: {self.data['Date'].min()} to {self.data['Date'].max()}")
        print(f"Unique tickers: {self.data['Ticker'].nunique()}")

        # Load S&P 500 index data for excess return calculations
        self._load_sp500_index_data()

    def _load_sp500_index_data(self):
        """
        Load S&P 500 index data (^GSPC) for calculating excess returns.
        """
        print("\nLoading S&P 500 index data for excess return calculations...")

        try:
            # Get date range from stock data
            start_date = self.data['Date'].min()
            end_date = self.data['Date'].max()

            # Download S&P 500 index data
            sp500_index = yf.download('^GSPC', start=start_date, end=end_date, progress=False)

            if sp500_index.empty:
                print("Warning: Could not download S&P 500 index data. Excess returns will not be calculated.")
                self.sp500_index = None
                return

            # Reset index to get Date as a column
            sp500_index = sp500_index.reset_index()

            # Rename columns if needed (yfinance returns 'Close' or might have multi-index)
            if isinstance(sp500_index.columns, pd.MultiIndex):
                sp500_index.columns = sp500_index.columns.get_level_values(0)

            # Keep only Date and Close
            self.sp500_index = sp500_index[['Date', 'Close']].copy()
            self.sp500_index.columns = ['Date', 'SP500_Close']

            # Ensure Date is datetime
            self.sp500_index['Date'] = pd.to_datetime(self.sp500_index['Date'], utc=True)
            self.sp500_index = self.sp500_index.sort_values('Date').reset_index(drop=True)

            print(f"Loaded S&P 500 index data: {len(self.sp500_index)} rows")
            print(f"S&P 500 date range: {self.sp500_index['Date'].min()} to {self.sp500_index['Date'].max()}")

        except Exception as e:
            print(f"Error loading S&P 500 index data: {e}")
            print("Excess returns will not be calculated.")
            self.sp500_index = None

    def calculate_rolling_entry_price(self):
        """
        Calculate rolling 5th percentile as realistic entry price.
        """
        print("\nCalculating rolling entry prices (5th percentile over 252 days)...")
        
        def calc_rolling_percentile(group):
            group = group.sort_values('Date')
            group['rolling_min_5pct'] = group['Close'].rolling(
                window=self.rolling_window, 
                min_periods=self.rolling_window
            ).quantile(self.percentile / 100)
            return group
        
        self.data = self.data.groupby('Ticker').apply(calc_rolling_percentile).reset_index()
        
        # Drop rows where we don't have enough history for rolling calculation
        before_drop = len(self.data)
        self.data = self.data.dropna(subset=['rolling_min_5pct'])
        after_drop = len(self.data)
        
        print(f"Dropped {before_drop - after_drop:,} rows without sufficient history")
        print(f"Remaining rows: {after_drop:,}")
        
    def detect_multiple_events(self):
        """
        Detect when stocks cross multiple thresholds (2x, 3x, etc.) from below.
        """
        print("\nDetecting multiple crossing events...")
        
        events = []
        
        for ticker in tqdm(self.data['Ticker'].unique(), desc="Processing tickers"):
            ticker_data = self.data[self.data['Ticker'] == ticker].copy()
            ticker_data = ticker_data.sort_values('Date').reset_index(drop=True)
            
            # Calculate multiples
            ticker_data['multiple'] = ticker_data['Close'] / ticker_data['rolling_min_5pct']
            
            # Track last event date for cooldown
            last_event_dates = {m: None for m in self.multiples}
            
            for i in range(1, len(ticker_data)):
                current_row = ticker_data.iloc[i]
                previous_row = ticker_data.iloc[i-1]
                
                current_date = current_row['Date']
                current_multiple = current_row['multiple']
                previous_multiple = previous_row['multiple']
                
                # Check each multiple threshold
                for multiple in self.multiples:
                    # Crossing condition: previous < threshold, current >= threshold
                    if previous_multiple < multiple <= current_multiple:
                        
                        # Check cooldown
                        last_event = last_event_dates[multiple]
                        if last_event is not None:
                            days_since_last = (current_date - last_event).days
                            if days_since_last < self.cooldown_days:
                                continue  # Skip due to cooldown
                        
                        # Record event
                        events.append({
                            'ticker': ticker,
                            'event_date': current_date,
                            'multiple': multiple,
                            'entry_price': current_row['rolling_min_5pct'],
                            'event_price': current_row['Close'],
                            'actual_multiple': current_multiple,
                            'in_sp500': current_row['in_sp500']
                        })
                        
                        # Update last event date
                        last_event_dates[multiple] = current_date
        
        self.events = pd.DataFrame(events)
        
        if len(self.events) > 0:
            print(f"\nDetected {len(self.events):,} events across {self.events['ticker'].nunique()} tickers")
            print("\nEvents by multiple:")
            print(self.events['multiple'].value_counts().sort_index())
        else:
            print("\nWARNING: No events detected!")
            
        return self.events
    
    def calculate_forward_returns(self):
        """
        Calculate forward returns for each event over different time horizons.
        """
        print("\nCalculating forward returns for each event...")
        
        results = []
        
        for idx, event in tqdm(self.events.iterrows(), total=len(self.events), desc="Calculating returns"):
            ticker = event['ticker']
            event_date = event['event_date']
            event_price = event['event_price']
            
            # Get future price data for this ticker
            future_data = self.data[
                (self.data['Ticker'] == ticker) & 
                (self.data['Date'] > event_date)
            ].sort_values('Date').reset_index(drop=True)
            
            if len(future_data) == 0:
                continue
            
            result = {
                'ticker': ticker,
                'event_date': event_date,
                'multiple': event['multiple'],
                'event_price': event_price,
                'entry_price': event['entry_price'],
                'in_sp500': event['in_sp500']
            }

            # Get S&P 500 price at event date for excess return calculation
            sp500_event_price = None
            if self.sp500_index is not None:
                sp500_at_event = self.sp500_index[self.sp500_index['Date'] <= event_date]
                if len(sp500_at_event) > 0:
                    sp500_event_price = sp500_at_event.iloc[-1]['SP500_Close']

            # Calculate metrics for each follow-up period
            for period_name, period_days in self.followup_periods.items():
                # Get data for this period
                period_data = future_data[future_data.index < period_days]

                if len(period_data) == 0:
                    continue
                
                # Calculate returns
                final_price = period_data.iloc[-1]['Close']
                max_price = period_data['Close'].max()
                min_price = period_data['Close'].min()
                
                total_return = (final_price - event_price) / event_price

                # Calculate S&P 500 return and excess return for this period
                sp500_return = None
                excess_return = None
                if sp500_event_price is not None and self.sp500_index is not None:
                    # Get S&P 500 price at period end
                    period_end_date = period_data.iloc[-1]['Date']
                    sp500_at_period_end = self.sp500_index[self.sp500_index['Date'] <= period_end_date]
                    if len(sp500_at_period_end) > 0:
                        sp500_final_price = sp500_at_period_end.iloc[-1]['SP500_Close']
                        sp500_return = (sp500_final_price - sp500_event_price) / sp500_event_price
                        excess_return = total_return - sp500_return

                # Calculate Maximum Drawdown
                cumulative_returns = (period_data['Close'] / event_price) - 1
                running_max = cumulative_returns.cummax()
                drawdown = cumulative_returns - running_max
                max_drawdown = drawdown.min()
                
                # Calculate volatility (annualized)
                returns = period_data['Close'].pct_change().dropna()
                volatility = returns.std() * np.sqrt(252)  # Annualized
                
                # Store results
                result[f'{period_name}_return'] = total_return
                result[f'{period_name}_sp500_return'] = sp500_return
                result[f'{period_name}_excess_return'] = excess_return
                result[f'{period_name}_final_price'] = final_price
                result[f'{period_name}_max_price'] = max_price
                result[f'{period_name}_min_price'] = min_price
                result[f'{period_name}_mdd'] = max_drawdown
                result[f'{period_name}_volatility'] = volatility
                result[f'{period_name}_days_available'] = len(period_data)
            
            results.append(result)
        
        self.results = pd.DataFrame(results)
        print(f"\nCalculated forward returns for {len(self.results):,} events")
        
        return self.results
    
    def calculate_next_multiple_probability(self):
        """
        Calculate probability of reaching the next multiple.
        Also tracks MDD until next multiple is reached and final returns if not reached.
        """
        print("\nCalculating probability of reaching next multiple...")
        
        next_multiple_results = []
        
        for idx, event in tqdm(self.events.iterrows(), total=len(self.events), desc="Calculating probabilities"):
            ticker = event['ticker']
            event_date = event['event_date']
            current_multiple = event['multiple']
            entry_price = event['entry_price']
            event_price = event['event_price']
            
            # Get future price data
            future_data = self.data[
                (self.data['Ticker'] == ticker) & 
                (self.data['Date'] > event_date)
            ].sort_values('Date').reset_index(drop=True)
            
            if len(future_data) == 0:
                continue
            
            # Check each follow-up period
            for period_name, period_days in self.followup_periods.items():
                period_data = future_data[future_data.index < period_days]

                if len(period_data) == 0:
                    continue

                # Check if next multiple was reached
                for next_mult in range(current_multiple + 1, 11):
                    target_price = entry_price * next_mult
                    reached = (period_data['Close'] >= target_price).any()
                    
                    if reached:
                        # Find when it was reached
                        first_reach_idx = period_data[period_data['Close'] >= target_price].index[0]
                        days_to_reach = first_reach_idx
                        
                        # Calculate MDD from event to when target was reached
                        data_until_reach = period_data.iloc[:first_reach_idx + 1]
                        cumulative_returns = (data_until_reach['Close'] / event_price) - 1
                        running_max = cumulative_returns.cummax()
                        drawdown = cumulative_returns - running_max
                        mdd_until_reach = drawdown.min()
                        
                        final_return_if_not_reached = np.nan
                    else:
                        days_to_reach = np.nan
                        mdd_until_reach = np.nan
                        
                        # Calculate final return if target was NOT reached
                        final_price = period_data.iloc[-1]['Close']
                        final_return_if_not_reached = (final_price - event_price) / event_price
                    
                    next_multiple_results.append({
                        'ticker': ticker,
                        'event_date': event_date,
                        'current_multiple': current_multiple,
                        'next_multiple': next_mult,
                        'period': period_name,
                        'reached': reached,
                        'days_to_reach': days_to_reach,
                        'mdd_until_reach': mdd_until_reach,
                        'final_return_if_not_reached': final_return_if_not_reached
                    })
        
        self.next_multiple_results = pd.DataFrame(next_multiple_results)
        print(f"Calculated {len(self.next_multiple_results):,} next-multiple checks")

        return self.next_multiple_results

    def calculate_multiple_distribution(self):
        """
        Calculate the distribution of multiples at the end of each period.
        Categorizes events into buckets: 0x-1x, 1x-2x, 2x-3x, 3x-4x, 4x-5x, 5x-10x, >10x.
        """
        print("\nCalculating multiple distribution at period ends...")

        distribution_results = []

        for idx, row in tqdm(self.results.iterrows(), total=len(self.results), desc="Calculating distributions"):
            ticker = row['ticker']
            event_date = row['event_date']
            current_multiple = row['multiple']
            entry_price = row['entry_price']

            for period_name in self.followup_periods.keys():
                # Check if data exists for this period
                return_col = f'{period_name}_return'
                if return_col not in row or pd.isna(row[return_col]):
                    continue

                final_price = row[f'{period_name}_final_price']

                # Calculate the multiple from entry price at period end
                final_multiple = final_price / entry_price if entry_price > 0 else 0

                # Categorize into buckets
                if final_multiple < 1:
                    bucket = '0x-1x'
                elif final_multiple < 2:
                    bucket = '1x-2x'
                elif final_multiple < 3:
                    bucket = '2x-3x'
                elif final_multiple < 4:
                    bucket = '3x-4x'
                elif final_multiple < 5:
                    bucket = '4x-5x'
                elif final_multiple < 10:
                    bucket = '5x-10x'
                else:
                    bucket = '>10x'

                distribution_results.append({
                    'ticker': ticker,
                    'event_date': event_date,
                    'current_multiple': current_multiple,
                    'period': period_name,
                    'entry_price': entry_price,
                    'final_price': final_price,
                    'final_multiple': final_multiple,
                    'bucket': bucket
                })

        self.multiple_distribution = pd.DataFrame(distribution_results)
        print(f"Calculated multiple distribution for {len(self.multiple_distribution):,} event-period combinations")

        return self.multiple_distribution

    def create_summary_statistics(self):
        """
        Create summary statistics table with all KPIs.
        """
        print("\nCreating summary statistics...")
        
        summary_rows = []
        
        for multiple in self.multiples:
            multiple_events = self.results[self.results['multiple'] == multiple]
            
            if len(multiple_events) == 0:
                continue
            
            row = {
                'multiple': f"{multiple}x",
                'n_events': len(multiple_events)
            }
            
            # Calculate statistics for each period
            for period_name in self.followup_periods.keys():
                returns = multiple_events[f'{period_name}_return'].dropna()
                mdds = multiple_events[f'{period_name}_mdd'].dropna()
                
                if len(returns) > 0:
                    # Return percentiles
                    row[f'{period_name}_return_25pct'] = returns.quantile(0.25)
                    row[f'{period_name}_return_50pct'] = returns.quantile(0.50)
                    row[f'{period_name}_return_75pct'] = returns.quantile(0.75)
                    row[f'{period_name}_return_mean'] = returns.mean()
                    
                    # MDD statistics
                    row[f'{period_name}_mdd_mean'] = mdds.mean()
                    row[f'{period_name}_mdd_median'] = mdds.median()
                    row[f'{period_name}_mdd_worst'] = mdds.min()
                    
                    # Win rate (positive returns)
                    row[f'{period_name}_win_rate'] = (returns > 0).mean()
                    
                    # Sharpe ratio (simplified: mean return / volatility)
                    vols = multiple_events[f'{period_name}_volatility'].dropna()
                    if len(vols) > 0:
                        sharpe = (returns.mean() / vols.mean()) if vols.mean() > 0 else np.nan
                        row[f'{period_name}_sharpe'] = sharpe
                    
                    # Sample size
                    row[f'{period_name}_n_samples'] = len(returns)
                # Probability of reaching higher multiples
                if hasattr(self, 'next_multiple_results'):
                    # --- KPIs for NEXT multiple (current + 1) ---
                    next_mult_data = self.next_multiple_results[
                        (self.next_multiple_results['current_multiple'] == multiple) &
                        (self.next_multiple_results['next_multiple'] == multiple + 1) &
                        (self.next_multiple_results['period'] == period_name)
                    ]

                    if len(next_mult_data) > 0:
                        prob = next_mult_data['reached'].mean()
                        row[f'{period_name}_prob_next_multiple'] = prob

                        # Average days to reach NEXT (for those that reached)
                        reached_data = next_mult_data[next_mult_data['reached']]
                        if len(reached_data) > 0:
                            avg_days = reached_data['days_to_reach'].mean()
                            row[f'{period_name}_avg_days_to_next'] = avg_days

                            # MDD until NEXT multiple is reached (percentiles)
                            mdd_until_next = reached_data['mdd_until_reach'].dropna()
                            if len(mdd_until_next) > 0:
                                row[f'{period_name}_mdd_until_next_25pct'] = mdd_until_next.quantile(0.25)
                                row[f'{period_name}_mdd_until_next_50pct'] = mdd_until_next.quantile(0.50)
                                row[f'{period_name}_mdd_until_next_75pct'] = mdd_until_next.quantile(0.75)

                        # Returns if NEXT multiple is NOT reached (percentiles)
                        not_reached_data = next_mult_data[~next_mult_data['reached']]
                        if len(not_reached_data) > 0:
                            returns_not_reached = not_reached_data['final_return_if_not_reached'].dropna()
                            if len(returns_not_reached) > 0:
                                row[f'{period_name}_return_not_next_25pct'] = returns_not_reached.quantile(0.25)
                                row[f'{period_name}_return_not_next_50pct'] = returns_not_reached.quantile(0.50)
                                row[f'{period_name}_return_not_next_75pct'] = returns_not_reached.quantile(0.75)

                    # --- Probabilities for ALL higher multiples ---
                    for target_mult in range(multiple + 1, 11):
                        target_data = self.next_multiple_results[
                            (self.next_multiple_results['current_multiple'] == multiple) &
                            (self.next_multiple_results['next_multiple'] == target_mult) &
                            (self.next_multiple_results['period'] == period_name)
                        ]

                        if len(target_data) > 0:
                            row[f'{period_name}_prob_reach_{target_mult}x'] = target_data['reached'].mean()
            
            summary_rows.append(row)
        
        self.summary = pd.DataFrame(summary_rows)
        print(f"Created summary table with {len(self.summary)} rows")

        return self.summary

    def create_kpi_tables(self):
        """
        Create new KPI tables with structured blocks for each multiple category.
        Creates one table per multiple (2x, 3x, 4x, 5x, 10x) with:
        - KPI Block 1: Total Return (percentiles, CAGR, % > 0)
        - KPI Block 2: Excess Return vs S&P 500 (percentiles, CAGR, % > 0)
        - KPI Block 3: Multiple Distribution (bucket percentages with N)
        """
        print("\nCreating new KPI tables...")

        # Only use 1Y, 2Y, 3Y, 5Y (not 10Y)
        periods = ['1Y', '2Y', '3Y', '5Y']
        # Focus on key multiples: 2x, 3x, 4x, 5x, 10x
        target_multiples = [2, 3, 4, 5, 10]

        kpi_tables = {}

        for multiple in target_multiples:
            # Filter events for this multiple
            multiple_events = self.results[self.results['multiple'] == multiple]

            if len(multiple_events) == 0:
                continue

            # Dictionary to store KPIs for this multiple
            kpis = {
                'multiple': f"{multiple}x",
                'n_events': len(multiple_events)
            }

            for period in periods:
                years = int(period.replace('Y', ''))

                # --- KPI Block 1: Total Return ---
                returns = multiple_events[f'{period}_return'].dropna()

                if len(returns) > 0:
                    # Percentiles
                    kpis[f'{period}_total_return_25pct'] = returns.quantile(0.25)
                    kpis[f'{period}_total_return_50pct'] = returns.quantile(0.50)
                    kpis[f'{period}_total_return_75pct'] = returns.quantile(0.75)

                    # CAGRs
                    kpis[f'{period}_total_cagr_25pct'] = self._compute_cagr_value(returns.quantile(0.25), years)
                    kpis[f'{period}_total_cagr_50pct'] = self._compute_cagr_value(returns.quantile(0.50), years)
                    kpis[f'{period}_total_cagr_75pct'] = self._compute_cagr_value(returns.quantile(0.75), years)

                    # % with return > 0
                    kpis[f'{period}_total_return_positive_pct'] = (returns > 0).mean() * 100

                # --- KPI Block 2: Excess Return vs S&P 500 ---
                excess_returns = multiple_events[f'{period}_excess_return'].dropna()

                if len(excess_returns) > 0:
                    # Percentiles
                    kpis[f'{period}_excess_return_25pct'] = excess_returns.quantile(0.25)
                    kpis[f'{period}_excess_return_50pct'] = excess_returns.quantile(0.50)
                    kpis[f'{period}_excess_return_75pct'] = excess_returns.quantile(0.75)

                    # CAGRs
                    kpis[f'{period}_excess_cagr_25pct'] = self._compute_cagr_value(excess_returns.quantile(0.25), years)
                    kpis[f'{period}_excess_cagr_50pct'] = self._compute_cagr_value(excess_returns.quantile(0.50), years)
                    kpis[f'{period}_excess_cagr_75pct'] = self._compute_cagr_value(excess_returns.quantile(0.75), years)

                    # % with excess return > 0
                    kpis[f'{period}_excess_return_positive_pct'] = (excess_returns > 0).mean() * 100

                # --- KPI Block 3: Multiple Distribution ---
                # Filter distribution data for this multiple and period
                dist_data = self.multiple_distribution[
                    (self.multiple_distribution['current_multiple'] == multiple) &
                    (self.multiple_distribution['period'] == period)
                ]

                if len(dist_data) > 0:
                    total_count = len(dist_data)
                    buckets = ['0x-1x', '1x-2x', '2x-3x', '3x-4x', '4x-5x', '5x-10x', '>10x']

                    for bucket in buckets:
                        bucket_count = len(dist_data[dist_data['bucket'] == bucket])
                        bucket_pct = (bucket_count / total_count * 100) if total_count > 0 else 0
                        kpis[f'{period}_dist_{bucket}_pct'] = bucket_pct
                        kpis[f'{period}_dist_{bucket}_n'] = bucket_count

            kpi_tables[f'{multiple}x'] = kpis

        self.kpi_tables = kpi_tables
        print(f"Created KPI tables for {len(kpi_tables)} multiples")

        return kpi_tables

    def _compute_cagr_value(self, total_return, years):
        """
        Compute CAGR as a decimal value (not formatted string).

        Parameters:
        -----------
        total_return : float
            Total return as decimal (e.g., 0.5 for 50%)
        years : int
            Number of years

        Returns:
        --------
        float : CAGR value as decimal, or None if invalid
        """
        if pd.isna(total_return) or years <= 0:
            return None

        growth = 1 + total_return
        if growth <= 0:
            return None

        cagr = growth ** (1 / years) - 1
        return cagr

    def _create_summary_html(self, output_path):
        """
        Create an HTML table from summary statistics with nice formatting.
        Uses external HTML template for better maintainability.
        
        Parameters:
        -----------
        output_path : Path
            Directory to save the HTML file
        """
        if not hasattr(self, 'summary') or self.summary is None:
            return
        
        # Load HTML template
        template_path = Path(__file__).parent / 'summary_template.html'
        
        if not template_path.exists():
            print(f"Warning: HTML template not found at {template_path}")
            print("Skipping HTML generation.")
            return
        
        with open(template_path, 'r', encoding='utf-8') as f:
            template = f.read()
        
        # Prepare data
        df = self.summary.copy()
        
        # Define column order and their metric types
        metric_config = [
            ('prob_higher', 'Prob Higher Multiples', 'prob_higher'),
            ('avg_days_to_next', 'Avg Days Next', 'days'),
            ('mdd_until_next', 'MDD Until Next', 'mdd_next'),
            ('return_not_next', 'Return Not Next', 'return_not_next'),
            ('return', 'Return Range', 'return'),
            ('implied_cagr', 'Implied CAGRs', 'implied_cagr'),
        ]
        
        # Group columns by period
        periods = ['1Y', '2Y', '3Y', '5Y', '10Y']
        
        # Build period headers with data-period attribute
        period_headers = ''
        for period in periods:
            col_count = len(metric_config)
            period_headers += f'<th colspan="{col_count}" data-period="{period}">{period}</th>\n'
        
        # Build metric headers with data-period-metric attribute
        metric_headers = ''
        for period in periods:
            for metric_key, display_name, metric_type in metric_config:
                metric_headers += f'<th data-period-metric="{period}-{metric_type}">{display_name}</th>\n'
        
        # Build table rows with data-period-metric attribute
        table_rows = ''
        for _, row in df.iterrows():
            table_rows += '                <tr>\n'
            table_rows += f'                    <td class="multiple-col">{row["multiple"]}</td>\n'
            table_rows += f'                    <td>{row["n_events"]}</td>\n'
            
            for period in periods:
                for metric_key, display_name, metric_type in metric_config:
                    value = self._format_cell_value(row, period, metric_key)
                    table_rows += f'                    <td data-period-metric="{period}-{metric_type}">{value}</td>\n'
            
            table_rows += '                </tr>\n'
        
        # Replace placeholders in template
        html = template.replace('{{PERIOD_HEADERS}}', period_headers)
        html = html.replace('{{METRIC_HEADERS}}', metric_headers)
        html = html.replace('{{TABLE_ROWS}}', table_rows)
        
        # Replace configuration placeholders
        html = html.replace('{{ROLLING_WINDOW}}', str(self.rolling_window))
        html = html.replace('{{PERCENTILE}}', str(self.percentile))
        html = html.replace('{{COOLDOWN_DAYS}}', str(self.cooldown_days))
        html = html.replace('{{MULTIPLES}}', ', '.join([f'{m}x' for m in self.multiples]))
        
        # Get date range from original data
        date_range_start = self.data['Date'].min().strftime('%Y-%m-%d')
        date_range_end = self.data['Date'].max().strftime('%Y-%m-%d')
        html = html.replace('{{DATE_RANGE_START}}', date_range_start)
        html = html.replace('{{DATE_RANGE_END}}', date_range_end)
        
        # Get total events count
        total_events = len(self.events) if hasattr(self, 'events') else 0
        html = html.replace('{{TOTAL_EVENTS}}', f'{total_events:,}')
        
        # Save HTML file
        with open(output_path / 'summary_statistics.html', 'w', encoding='utf-8') as f:
            f.write(html)

    def _create_kpi_tables_html(self, output_path):
        """
        Create HTML output for the new KPI tables.
        One table per multiple category (2x, 3x, 4x, 5x, 10x).
        """
        if not hasattr(self, 'kpi_tables') or not self.kpi_tables:
            print("Warning: No KPI tables to generate HTML from.")
            return

        print("\nGenerating HTML for KPI tables...")

        # Load HTML template
        template_path = Path(__file__).parent / 'kpi_tables_template.html'

        if not template_path.exists():
            print(f"Warning: KPI tables template not found at {template_path}")
            print("Creating simple HTML output instead.")
            self._create_simple_kpi_html(output_path)
            return

        with open(template_path, 'r', encoding='utf-8') as f:
            template = f.read()

        # Generate HTML for all tables
        periods = ['1Y', '2Y', '3Y', '5Y']
        all_tables_html = ''

        for multiple_key in ['2x', '3x', '4x', '5x', '10x']:
            if multiple_key not in self.kpi_tables:
                continue

            kpis = self.kpi_tables[multiple_key]
            n_events = kpis.get('n_events', 0)

            # Build table HTML for this multiple
            table_html = f'''
            <div class="kpi-table-container">
                <h2>{multiple_key} Events (N={n_events})</h2>
                <table class="kpi-table">
                    <thead>
                        <tr>
                            <th class="kpi-name">KPI</th>
'''

            # Add period headers
            for period in periods:
                table_html += f'                            <th>{period}</th>\n'

            table_html += '''                        </tr>
                    </thead>
                    <tbody>
'''

            # --- KPI Block 1: Total Return ---
            table_html += '                        <tr class="block-header"><td colspan="5"><strong>Total Return ab Event Date</strong></td></tr>\n'

            # Row 1: Total Return Percentiles
            table_html += '                        <tr>\n'
            table_html += '                            <td class="kpi-name">Total Return (25%-50%-75%)</td>\n'
            for period in periods:
                p25 = kpis.get(f'{period}_total_return_25pct')
                p50 = kpis.get(f'{period}_total_return_50pct')
                p75 = kpis.get(f'{period}_total_return_75pct')
                if p25 is not None and p50 is not None and p75 is not None:
                    table_html += f'                            <td><span class="range-small">{p25*100:.1f}%</span> - <span class="range-mid">{p50*100:.1f}%</span> - <span class="range-small">{p75*100:.1f}%</span></td>\n'
                else:
                    table_html += '                            <td>N/A</td>\n'
            table_html += '                        </tr>\n'

            # Row 2: CAGR
            table_html += '                        <tr>\n'
            table_html += '                            <td class="kpi-name">CAGR (25%-50%-75%)</td>\n'
            for period in periods:
                c25 = kpis.get(f'{period}_total_cagr_25pct')
                c50 = kpis.get(f'{period}_total_cagr_50pct')
                c75 = kpis.get(f'{period}_total_cagr_75pct')
                if c25 is not None and c50 is not None and c75 is not None:
                    table_html += f'                            <td><span class="range-small">{c25*100:.1f}%</span> - <span class="range-mid">{c50*100:.1f}%</span> - <span class="range-small">{c75*100:.1f}%</span></td>\n'
                else:
                    table_html += '                            <td>N/A</td>\n'
            table_html += '                        </tr>\n'

            # Row 3: % with Return > 0
            table_html += '                        <tr>\n'
            table_html += '                            <td class="kpi-name">% Events mit Total Return > 0%</td>\n'
            for period in periods:
                pct = kpis.get(f'{period}_total_return_positive_pct')
                if pct is not None:
                    table_html += f'                            <td>{pct:.1f}%</td>\n'
                else:
                    table_html += '                            <td>N/A</td>\n'
            table_html += '                        </tr>\n'

            # --- KPI Block 2: Excess Return ---
            table_html += '                        <tr class="block-header"><td colspan="5"><strong>Excess Return vs. S&P 500 ab Event Date</strong></td></tr>\n'

            # Row 4: Excess Return Percentiles
            table_html += '                        <tr>\n'
            table_html += '                            <td class="kpi-name">Excess Return (25%-50%-75%)</td>\n'
            for period in periods:
                p25 = kpis.get(f'{period}_excess_return_25pct')
                p50 = kpis.get(f'{period}_excess_return_50pct')
                p75 = kpis.get(f'{period}_excess_return_75pct')
                if p25 is not None and p50 is not None and p75 is not None:
                    table_html += f'                            <td><span class="range-small">{p25*100:.1f}%</span> - <span class="range-mid">{p50*100:.1f}%</span> - <span class="range-small">{p75*100:.1f}%</span></td>\n'
                else:
                    table_html += '                            <td>N/A</td>\n'
            table_html += '                        </tr>\n'

            # Row 5: Excess CAGR
            table_html += '                        <tr>\n'
            table_html += '                            <td class="kpi-name">Excess CAGR (25%-50%-75%)</td>\n'
            for period in periods:
                c25 = kpis.get(f'{period}_excess_cagr_25pct')
                c50 = kpis.get(f'{period}_excess_cagr_50pct')
                c75 = kpis.get(f'{period}_excess_cagr_75pct')
                if c25 is not None and c50 is not None and c75 is not None:
                    table_html += f'                            <td><span class="range-small">{c25*100:.1f}%</span> - <span class="range-mid">{c50*100:.1f}%</span> - <span class="range-small">{c75*100:.1f}%</span></td>\n'
                else:
                    table_html += '                            <td>N/A</td>\n'
            table_html += '                        </tr>\n'

            # Row 6: % with Excess Return > 0
            table_html += '                        <tr>\n'
            table_html += '                            <td class="kpi-name">% Events mit Excess Return > 0%</td>\n'
            for period in periods:
                pct = kpis.get(f'{period}_excess_return_positive_pct')
                if pct is not None:
                    table_html += f'                            <td>{pct:.1f}%</td>\n'
                else:
                    table_html += '                            <td>N/A</td>\n'
            table_html += '                        </tr>\n'

            # --- KPI Block 3: Multiple Distribution ---
            table_html += '                        <tr class="block-header"><td colspan="5"><strong>Multiple-Verteilung am Periodenende</strong></td></tr>\n'

            buckets = ['0x-1x', '1x-2x', '2x-3x', '3x-4x', '4x-5x', '5x-10x', '>10x']
            for bucket in buckets:
                table_html += '                        <tr>\n'
                table_html += f'                            <td class="kpi-name">{bucket}</td>\n'
                for period in periods:
                    pct = kpis.get(f'{period}_dist_{bucket}_pct')
                    n = kpis.get(f'{period}_dist_{bucket}_n', 0)
                    if pct is not None:
                        table_html += f'                            <td>{pct:.1f}%<br><span class="bucket-count">N={n}</span></td>\n'
                    else:
                        table_html += '                            <td>N/A</td>\n'
                table_html += '                        </tr>\n'

            table_html += '''                    </tbody>
                </table>
            </div>
'''

            all_tables_html += table_html

        # Replace placeholder in template
        html = template.replace('{{TABLES}}', all_tables_html)

        # Replace configuration placeholders
        html = html.replace('{{ROLLING_WINDOW}}', str(self.rolling_window))
        html = html.replace('{{PERCENTILE}}', str(self.percentile))
        html = html.replace('{{COOLDOWN_DAYS}}', str(self.cooldown_days))

        # Get date range
        date_range_start = self.data['Date'].min().strftime('%Y-%m-%d')
        date_range_end = self.data['Date'].max().strftime('%Y-%m-%d')
        html = html.replace('{{DATE_RANGE_START}}', date_range_start)
        html = html.replace('{{DATE_RANGE_END}}', date_range_end)

        # Get total events
        total_events = len(self.events) if hasattr(self, 'events') else 0
        html = html.replace('{{TOTAL_EVENTS}}', f'{total_events:,}')

        # Save HTML file
        output_file = output_path / 'kpi_tables.html'
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html)

        print(f"Saved KPI tables HTML to {output_file}")

    def _create_simple_kpi_html(self, output_path):
        """
        Create a simple HTML output for KPI tables if template is not available.
        """
        print("Creating simple KPI HTML output...")

        html = '''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Event Study KPI Tables</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        .kpi-table-container { margin-bottom: 40px; }
        .kpi-table { border-collapse: collapse; width: 100%; margin-top: 10px; }
        .kpi-table th, .kpi-table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        .kpi-table th { background-color: #4CAF50; color: white; }
        .kpi-table .block-header { background-color: #e0e0e0; font-weight: bold; }
        .kpi-name { font-weight: bold; }
        .range-small { font-size: 0.9em; color: #666; }
        .range-mid { font-weight: bold; }
        .bucket-count { font-size: 0.8em; color: #666; }
    </style>
</head>
<body>
    <h1>Event Study KPI Tables</h1>
'''

        # Generate simple tables
        periods = ['1Y', '2Y', '3Y', '5Y']

        for multiple_key in ['2x', '3x', '4x', '5x', '10x']:
            if multiple_key not in self.kpi_tables:
                continue

            kpis = self.kpi_tables[multiple_key]
            n_events = kpis.get('n_events', 0)

            html += f'<div class="kpi-table-container"><h2>{multiple_key} Events (N={n_events})</h2>'
            html += '<table class="kpi-table"><thead><tr><th class="kpi-name">KPI</th>'

            for period in periods:
                html += f'<th>{period}</th>'

            html += '</tr></thead><tbody>'

            # Add all KPI rows (simplified version)
            html += '<tr class="block-header"><td colspan="5"><strong>Total Return ab Event Date</strong></td></tr>'
            html += '<tr><td class="kpi-name">Total Return (25%-50%-75%)</td>'
            for period in periods:
                p25 = kpis.get(f'{period}_total_return_25pct')
                p50 = kpis.get(f'{period}_total_return_50pct')
                p75 = kpis.get(f'{period}_total_return_75pct')
                if p25 is not None and p50 is not None and p75 is not None:
                    html += f'<td><span class="range-small">{p25*100:.1f}%</span> - <span class="range-mid">{p50*100:.1f}%</span> - <span class="range-small">{p75*100:.1f}%</span></td>'
                else:
                    html += '<td>N/A</td>'
            html += '</tr>'

            html += '</tbody></table></div>'

        html += '</body></html>'

        output_file = output_path / 'kpi_tables.html'
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html)

        print(f"Saved simple KPI tables HTML to {output_file}")

    def _format_cell_value(self, row, period, metric):
        """
        Format a single cell value, combining ranges where applicable.
        
        Parameters:
        -----------
        row : pd.Series
            Row from summary dataframe
        period : str
            Period (1Y, 2Y, etc.)
        metric : str
            Metric name
            
        Returns:
        --------
        str : Formatted HTML value
        """
        if metric == 'prob_higher':
            # Combined display of probabilities for ALL higher multiples
            # Extract the current multiple number from the row
            mult_str = row.get('multiple', '')
            if isinstance(mult_str, str) and mult_str.endswith('x'):
                current_mult = int(mult_str[:-1])
            else:
                return 'N/A'

            parts = []
            for target in range(current_mult + 1, 11):
                col = f'{period}_prob_reach_{target}x'
                if col in row and pd.notna(row[col]):
                    prob_pct = row[col] * 100
                    parts.append(f'<span class="prob-item">{target}x: {prob_pct:.0f}%</span>')

            if parts:
                return '<span class="prob-higher-list">' + ' '.join(parts) + '</span>'
            return 'N/A'

        elif metric == 'avg_days_to_next':
            col = f'{period}_avg_days_to_next'
            if col in row and pd.notna(row[col]):
                return f'{row[col]:.0f}'
            return 'N/A'

        # Range metrics (25%, 50%, 75% combined)
        elif metric == 'return':
            col_25 = f'{period}_return_25pct'
            col_50 = f'{period}_return_50pct'
            col_75 = f'{period}_return_75pct'
            years = int(period.replace('Y', ''))
            return self._format_range(row, col_25, col_50, col_75, is_percentage=True, cagr_years=years)

        elif metric == 'mdd_until_next':
            col_25 = f'{period}_mdd_until_next_25pct'
            col_50 = f'{period}_mdd_until_next_50pct'
            col_75 = f'{period}_mdd_until_next_75pct'
            return self._format_range(row, col_25, col_50, col_75, is_percentage=True)

        elif metric == 'return_not_next':
            col_25 = f'{period}_return_not_next_25pct'
            col_50 = f'{period}_return_not_next_50pct'
            col_75 = f'{period}_return_not_next_75pct'
            return self._format_range(row, col_25, col_50, col_75, is_percentage=True)

        elif metric == 'implied_cagr':
            return self._format_implied_cagr(row, period)

        return 'N/A'
    
    def _format_range(self, row, col_25, col_50, col_75, is_percentage=True, cagr_years=None):
        """
        Format a range value in the format: small - BOLD - small
        Optionally adds a second CAGR line if cagr_years is provided.

        Parameters:
        -----------
        row : pd.Series
            Row from dataframe
        col_25, col_50, col_75 : str
            Column names for 25th, 50th, 75th percentiles
        is_percentage : bool
            Whether to format as percentage
        cagr_years : int or None
            If provided, adds a second line with annualized CAGR values

        Returns:
        --------
        str : Formatted HTML string
        """
        if col_25 not in row or col_50 not in row or col_75 not in row:
            return 'N/A'

        val_25 = row[col_25]
        val_50 = row[col_50]
        val_75 = row[col_75]

        # Check if all values are available
        if pd.isna(val_25) or pd.isna(val_50) or pd.isna(val_75):
            return 'N/A'

        # Format values
        if is_percentage:
            str_25 = f'{val_25*100:.1f}%'
            str_50 = f'{val_50*100:.1f}%'
            str_75 = f'{val_75*100:.1f}%'
        else:
            str_25 = f'{val_25:.1f}'
            str_50 = f'{val_50:.1f}'
            str_75 = f'{val_75:.1f}'

        # Build HTML with styling
        html = f'<span class="range-value">'
        html += f'<span class="small">{str_25}</span>'
        html += f'<span class="mid"> {str_50} </span>'
        html += f'<span class="small">{str_75}</span>'
        html += '</span>'

        # Add CAGR second line if requested
        if cagr_years is not None and cagr_years > 0:
            cagr_25 = self._compute_cagr(val_25, cagr_years)
            cagr_50 = self._compute_cagr(val_50, cagr_years)
            cagr_75 = self._compute_cagr(val_75, cagr_years)
            html += '<br>'
            html += f'<span class="range-value cagr-line">'
            html += f'<span class="small">{cagr_25}</span>'
            html += f'<span class="mid"> {cagr_50} </span>'
            html += f'<span class="small">{cagr_75}</span>'
            html += '</span>'

        return html

    def _compute_cagr(self, total_return, years):
        """
        Compute annualized CAGR from a total return over a number of years.

        Parameters:
        -----------
        total_return : float
            Total return as decimal (e.g., 0.5 for 50%)
        years : int
            Number of years

        Returns:
        --------
        str : Formatted CAGR string
        """
        growth = 1 + total_return
        if growth <= 0:
            return 'N/A'
        cagr = growth ** (1 / years) - 1
        return f'{cagr*100:.1f}%'

    def _format_implied_cagr(self, row, period):
        """
        Format implied CAGR for reaching each higher multiple within the given period.
        Shows what annualized return is needed from current multiple to each higher multiple.

        Parameters:
        -----------
        row : pd.Series
            Row from summary dataframe
        period : str
            Period (1Y, 2Y, etc.)

        Returns:
        --------
        str : Formatted HTML string
        """
        mult_str = row.get('multiple', '')
        if isinstance(mult_str, str) and mult_str.endswith('x'):
            current_mult = int(mult_str[:-1])
        else:
            return 'N/A'

        years = int(period.replace('Y', ''))

        parts = []
        for target in range(current_mult + 1, 11):
            # Growth factor from current multiple to target multiple
            growth = target / current_mult
            cagr = growth ** (1 / years) - 1
            parts.append(f'<span class="prob-item">{target}x: {cagr*100:.1f}%</span>')

        if parts:
            return '<span class="prob-higher-list">' + ' '.join(parts) + '</span>'
        return 'N/A'
    
    def export_ticker_to_excel(self, ticker, output_dir='analysis_results'):
        """
        Export all analysis data for a single ticker to Excel for manual verification.
        
        Parameters:
        -----------
        ticker : str
            The ticker symbol to export
        output_dir : str
            Directory to save the Excel file
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        print(f"\nExporting data for {ticker} to Excel...")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Price History with rolling calculations
        ws1 = wb.create_sheet("Price History")
        ticker_data = self.data[self.data['Ticker'] == ticker].copy()
        ticker_data = ticker_data.sort_values('Date').reset_index(drop=True)
        
        # Calculate multiple for visualization
        if 'rolling_min_5pct' in ticker_data.columns:
            ticker_data['multiple'] = ticker_data['Close'] / ticker_data['rolling_min_5pct']
        
        # Add headers
        headers = ['Date', 'Close', 'Rolling_5pct', 'Multiple', 'In_SP500']
        ws1.append(headers)
        
        # Style header
        for cell in ws1[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for _, row in ticker_data.iterrows():
            ws1.append([
                row['Date'].strftime('%Y-%m-%d') if pd.notna(row['Date']) else '',
                row['Close'] if pd.notna(row['Close']) else '',
                row.get('rolling_min_5pct', '') if pd.notna(row.get('rolling_min_5pct')) else '',
                row.get('multiple', '') if pd.notna(row.get('multiple')) else '',
                row.get('in_sp500', '')
            ])
        
        # Format columns
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 12
        
        # Sheet 2: Detected Events
        ws2 = wb.create_sheet("Events")
        if hasattr(self, 'events'):
            ticker_events = self.events[self.events['ticker'] == ticker].copy()
            
            headers = ['Event_Date', 'Multiple', 'Entry_Price', 'Event_Price', 'Actual_Multiple', 'In_SP500']
            ws2.append(headers)
            
            for cell in ws2[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            for _, row in ticker_events.iterrows():
                ws2.append([
                    row['event_date'].strftime('%Y-%m-%d'),
                    row['multiple'],
                    row['entry_price'],
                    row['event_price'],
                    row['actual_multiple'],
                    row['in_sp500']
                ])
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws2.column_dimensions[col].width = 15
        
        # Sheet 3: Forward Returns by Event
        ws3 = wb.create_sheet("Forward Returns")
        if hasattr(self, 'results'):
            ticker_results = self.results[self.results['ticker'] == ticker].copy()
            
            if len(ticker_results) > 0:
                # Write headers
                headers = ['Event_Date', 'Multiple'] + [
                    f'{period}_{metric}' 
                    for period in ['1Y', '2Y', '3Y', '5Y', '10Y']
                    for metric in ['return', 'final_price', 'max_price', 'mdd']
                ]
                ws3.append(headers)
                
                for cell in ws3[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data
                for _, row in ticker_results.iterrows():
                    data_row = [
                        row['event_date'].strftime('%Y-%m-%d'),
                        row['multiple']
                    ]
                    for period in ['1Y', '2Y', '3Y', '5Y', '10Y']:
                        data_row.extend([
                            row.get(f'{period}_return', ''),
                            row.get(f'{period}_final_price', ''),
                            row.get(f'{period}_max_price', ''),
                            row.get(f'{period}_mdd', '')
                        ])
                    ws3.append(data_row)
                
                # Auto-width for columns
                for col_idx in range(1, len(headers) + 1):
                    ws3.column_dimensions[chr(64 + col_idx)].width = 14
        
        # Sheet 4: Next Multiple Probabilities
        ws4 = wb.create_sheet("Next Multiple Analysis")
        if hasattr(self, 'next_multiple_results'):
            ticker_next = self.next_multiple_results[
                self.next_multiple_results['ticker'] == ticker
            ].copy()
            
            if len(ticker_next) > 0:
                headers = ['Event_Date', 'Current_Multiple', 'Next_Multiple', 'Period', 
                          'Reached', 'Days_to_Reach', 'MDD_Until_Reach', 'Return_If_Not_Reached']
                ws4.append(headers)
                
                for cell in ws4[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                for _, row in ticker_next.iterrows():
                    ws4.append([
                        row['event_date'].strftime('%Y-%m-%d'),
                        row['current_multiple'],
                        row['next_multiple'],
                        row['period'],
                        row['reached'],
                        row.get('days_to_reach', ''),
                        row.get('mdd_until_reach', ''),
                        row.get('final_return_if_not_reached', '')
                    ])
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws4.column_dimensions[col].width = 18
        
        # Save file
        filename = output_path / f'ticker_analysis_{ticker}.xlsx'
        wb.save(filename)
        
        print(f"Exported {ticker} analysis to {filename}")
        
        return filename
    
    def save_results(self, output_dir='analysis_results'):
        """
        Save all results to files.
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        print(f"\nSaving results to {output_dir}/...")

        # Save detailed results
        if hasattr(self, 'results'):
            self.results.to_csv(output_path / 'detailed_results.csv', index=False)
            print("Saved detailed_results.csv")

        # Save summary statistics
        if hasattr(self, 'summary'):
            self.summary.to_csv(output_path / 'summary_statistics.csv', index=False)
            print("Saved summary_statistics.csv")

            # Also save as HTML
            self._create_summary_html(output_path)
            print("Saved summary_statistics.html")

        # Save events
        if hasattr(self, 'events'):
            self.events.to_csv(output_path / 'detected_events.csv', index=False)
            print("Saved detected_events.csv")

        # Save next multiple probabilities
        if hasattr(self, 'next_multiple_results'):
            self.next_multiple_results.to_csv(output_path / 'next_multiple_probabilities.csv', index=False)
            print("Saved next_multiple_probabilities.csv")

        # Save multiple distribution
        if hasattr(self, 'multiple_distribution'):
            self.multiple_distribution.to_csv(output_path / 'multiple_distribution.csv', index=False)
            print("Saved multiple_distribution.csv")

        # Save KPI tables
        if hasattr(self, 'kpi_tables'):
            # Save as CSV
            kpi_df = pd.DataFrame(self.kpi_tables).T
            kpi_df.to_csv(output_path / 'kpi_tables.csv')
            print("Saved kpi_tables.csv")

            # Save as HTML
            self._create_kpi_tables_html(output_path)
            print("Saved kpi_tables.html")

        # Save run configuration
        run_config = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'rolling_window': self.rolling_window,
            'percentile': self.percentile,
            'cooldown_days': self.cooldown_days,
            'multiples': self.multiples,
            'followup_periods': self.followup_periods,
            'date_range': [
                self.data['Date'].min().strftime('%Y-%m-%d'),
                self.data['Date'].max().strftime('%Y-%m-%d')
            ],
            'unique_tickers': int(self.data['Ticker'].nunique()),
            'total_events': len(self.events) if hasattr(self, 'events') else 0,
            'total_results': len(self.results) if hasattr(self, 'results') else 0
        }

        with open(output_path / 'run_config.json', 'w') as f:
            json.dump(run_config, f, indent=2)
        print("Saved run_config.json")

        print(f"\nAll results saved to {output_dir}/")
    
    def run_full_analysis(self, output_dir='analysis_results'):
        """
        Run the complete analysis pipeline.
        """
        print("\n" + "="*70)
        print("WINNER STOCKS ANALYSIS")
        print("="*70)
        
        # Step 1: Calculate rolling entry prices
        self.calculate_rolling_entry_price()
        
        # Step 2: Detect multiple events
        self.detect_multiple_events()
        
        if len(self.events) == 0:
            print("\nERROR: No events detected. Cannot continue analysis.")
            return
        
        # Step 3: Calculate forward returns
        self.calculate_forward_returns()
        
        # Step 4: Calculate next multiple probabilities
        self.calculate_next_multiple_probability()

        # Step 5: Calculate multiple distribution
        self.calculate_multiple_distribution()

        # Step 6: Create summary statistics (old format)
        self.create_summary_statistics()

        # Step 7: Create new KPI tables (new format)
        self.create_kpi_tables()

        # Step 8: Save results
        self.save_results(output_dir)
        
        print("\n" + "="*70)
        print("ANALYSIS COMPLETE!")
        print("="*70)
        print(f"\nResults saved to: {output_dir}/")
        print("\nGenerated files:")
        print("  - kpi_tables.html (NEW: KPI tables with Total/Excess Return & Distribution)")
        print("  - kpi_tables.csv (NEW: KPI tables in CSV format)")
        print("  - summary_statistics.csv (main results table)")
        print("  - summary_statistics.html (interactive HTML table)")
        print("  - detailed_results.csv (all individual events)")
        print("  - detected_events.csv (all crossing events)")
        print("  - next_multiple_probabilities.csv")
        print("  - multiple_distribution.csv (NEW: multiple distribution data)")
        print("\nTo export data for a specific ticker to Excel:")
        print("  analyzer.export_ticker_to_excel('AAPL')")
        
        return self.summary


def list_runs(base_dir='analysis_results'):
    """
    List all saved analysis runs with their configurations.
    """
    base_path = Path(base_dir)

    if not base_path.exists():
        print(f"No analysis results found. Directory '{base_dir}/' does not exist.")
        return

    # Check for run_config.json in the base directory itself (non-versioned run)
    runs = []
    base_config = base_path / 'run_config.json'
    if base_config.exists():
        with open(base_config, 'r') as f:
            config = json.load(f)
        runs.append(('(default)', config))

    # Check subdirectories
    for subdir in sorted(base_path.iterdir()):
        if subdir.is_dir():
            config_file = subdir / 'run_config.json'
            if config_file.exists():
                with open(config_file, 'r') as f:
                    config = json.load(f)
                runs.append((subdir.name, config))

    if not runs:
        print("No saved runs found.")
        return

    print(f"\n{'='*90}")
    print(f"{'SAVED ANALYSIS RUNS':^90}")
    print(f"{'='*90}")
    print(f"\n{'Run Name':<20} {'Timestamp':<22} {'Window':<8} {'Pctl':<6} {'Cooldown':<10} {'Events':<8} {'Date Range'}")
    print(f"{'-'*20} {'-'*22} {'-'*8} {'-'*6} {'-'*10} {'-'*8} {'-'*23}")

    for name, config in runs:
        timestamp = config.get('timestamp', 'N/A')
        window = config.get('rolling_window', 'N/A')
        pctl = config.get('percentile', 'N/A')
        cooldown = config.get('cooldown_days', 'N/A')
        events = config.get('total_events', 'N/A')
        date_range = config.get('date_range', ['N/A', 'N/A'])
        date_str = f"{date_range[0]} to {date_range[1]}"

        print(f"{name:<20} {timestamp:<22} {str(window):<8} {str(pctl):<6} {str(cooldown):<10} {str(events):<8} {date_str}")

    print(f"\nTotal: {len(runs)} run(s) found in {base_dir}/")


def main():
    """
    Main function to run analysis with CLI argument support.
    """
    import argparse
    from download_sp500_data import SP500DataManager

    parser = argparse.ArgumentParser(
        description='Analyze winner stocks (Multibagger events) in the S&P 500.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python analyze_winner_stocks.py                                  # Default run
  python analyze_winner_stocks.py --run-name baseline              # Named run
  python analyze_winner_stocks.py --run-name cd60 --cooldown-days 60
  python analyze_winner_stocks.py --run-name p10 --percentile 10
  python analyze_winner_stocks.py --list-runs                      # Show all runs
        """
    )

    parser.add_argument('--run-name', type=str, default=None,
                        help='Name for this analysis run. Results are saved to analysis_results/<run-name>/. '
                             'If not provided, results are saved directly to analysis_results/.')
    parser.add_argument('--rolling-window', type=int, default=None,
                        help='Rolling window in trading days for entry price calculation (default: 252)')
    parser.add_argument('--percentile', type=int, default=None,
                        help='Percentile for rolling entry price (default: 5)')
    parser.add_argument('--cooldown-days', type=int, default=None,
                        help='Cooldown period in days between events (default: 90)')
    parser.add_argument('--list-runs', action='store_true',
                        help='List all saved analysis runs and exit')
    parser.add_argument('--export-ticker', type=str, default=None,
                        help='Export detailed Excel workbook for a specific ticker (e.g., AAPL)')

    args = parser.parse_args()

    # Handle --list-runs
    if args.list_runs:
        list_runs()
        return

    print("Loading S&P 500 historical data...")

    # Load data
    manager = SP500DataManager()
    df, metadata = manager.load_data()

    if df is None:
        print("ERROR: Could not load data. Please run download_sp500_data.py first.")
        return

    # Only analyze stocks that were in S&P 500
    print(f"\nFiltering for S&P 500 members only...")
    df_sp500 = df[df['in_sp500']].copy()
    print(f"Rows after filtering: {len(df_sp500):,} (from {len(df):,})")

    # Create analyzer
    analyzer = WinnerStocksAnalyzer(df_sp500)

    # Apply parameter overrides
    if args.rolling_window is not None:
        analyzer.rolling_window = args.rolling_window
        print(f"Override: rolling_window = {args.rolling_window}")
    if args.percentile is not None:
        analyzer.percentile = args.percentile
        print(f"Override: percentile = {args.percentile}")
    if args.cooldown_days is not None:
        analyzer.cooldown_days = args.cooldown_days
        print(f"Override: cooldown_days = {args.cooldown_days}")

    # Determine output directory
    base_dir = 'analysis_results'
    if args.run_name:
        output_dir = str(Path(base_dir) / args.run_name)
    else:
        output_dir = base_dir

    # Run analysis
    summary = analyzer.run_full_analysis(output_dir=output_dir)

    # Export ticker if requested
    if args.export_ticker:
        analyzer.export_ticker_to_excel(args.export_ticker, output_dir=output_dir)
    else:
        print("\n" + "="*70)
        print("EXAMPLE: Exporting single ticker data")
        print("="*70)
        if hasattr(analyzer, 'events') and len(analyzer.events) > 0:
            example_ticker = "AAPL"
            analyzer.export_ticker_to_excel(example_ticker, output_dir=output_dir)
            print(f"\nTo export another ticker, use:")
            print(f"  python analyze_winner_stocks.py --export-ticker YOUR_TICKER", end="")
            if args.run_name:
                print(f" --run-name {args.run_name}")
            else:
                print()


if __name__ == "__main__":
    main()
