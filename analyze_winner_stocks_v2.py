"""
Winner Stocks Analysis Script

Analyzes what happens AFTER a stock achieves 2x, 3x, 4x, ... 10x returns.
Answers the question: How long should you hold your winners?

Key Features:
- Uses rolling 5th percentile (252 days) as entry point
- Detects crossing events (from below to above threshold)
- 90-day cooldown period between events
- Tracks performance over 1, 2, 3, 5, and 10 years
- Comprehensive KPIs including returns, drawdowns, probabilities
"""

import pandas as pd
import numpy as np
from datetime import timedelta
import warnings
from tqdm import tqdm
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

# Set style for plots
sns.set_style("whitegrid")
plt.rcParams['figure.figsize'] = (14, 8)


class WinnerStocksAnalyzer:
    """
    Analyzer for winner stocks performance after achieving multiples.
    """
    
    def __init__(self, price_data):
        """
        Initialize analyzer with price data.
        
        Parameters:
        -----------
        price_data : pd.DataFrame
            DataFrame from download_sp500_data.py with columns:
            Date, Ticker, Close, in_sp500, etc.
        """
        self.data = price_data.copy()
        self.data = self.data.sort_values(['Ticker', 'Date']).reset_index(drop=True)
        
        # Configuration
        self.rolling_window = 252  # 1 year for rolling minimum
        self.percentile = 5  # 5th percentile
        self.cooldown_days = 90  # 90 days between events
        self.multiples = [2, 3, 4, 5, 6, 7, 8, 9, 10]  # Multiples to analyze
        self.followup_periods = {
            '1Y': 252,
            '2Y': 504,
            '3Y': 756,
            '5Y': 1260,
            '10Y': 2520
        }
        
        print("Initializing Winner Stocks Analyzer...")
        print(f"Data shape: {self.data.shape}")
        print(f"Date range: {self.data['Date'].min()} to {self.data['Date'].max()}")
        print(f"Unique tickers: {self.data['Ticker'].nunique()}")
        
    def calculate_rolling_entry_price(self):
        """
        Calculate rolling 5th percentile as realistic entry price.
        """
        print("\nCalculating rolling entry prices (5th percentile over 252 days)...")
        
        def calc_rolling_percentile(group):
            group = group.sort_values('Date')
            group['rolling_min_5pct'] = group['Close'].rolling(
                window=self.rolling_window, 
                min_periods=self.rolling_window
            ).quantile(self.percentile / 100)
            return group
        
        self.data = self.data.groupby('Ticker').apply(calc_rolling_percentile).reset_index()
        
        # Drop rows where we don't have enough history for rolling calculation
        before_drop = len(self.data)
        self.data = self.data.dropna(subset=['rolling_min_5pct'])
        after_drop = len(self.data)
        
        print(f"Dropped {before_drop - after_drop:,} rows without sufficient history")
        print(f"Remaining rows: {after_drop:,}")
        
    def detect_multiple_events(self):
        """
        Detect when stocks cross multiple thresholds (2x, 3x, etc.) from below.
        """
        print("\nDetecting multiple crossing events...")
        
        events = []
        
        for ticker in tqdm(self.data['Ticker'].unique(), desc="Processing tickers"):
            ticker_data = self.data[self.data['Ticker'] == ticker].copy()
            ticker_data = ticker_data.sort_values('Date').reset_index(drop=True)
            
            # Calculate multiples
            ticker_data['multiple'] = ticker_data['Close'] / ticker_data['rolling_min_5pct']
            
            # Track last event date for cooldown
            last_event_dates = {m: None for m in self.multiples}
            
            for i in range(1, len(ticker_data)):
                current_row = ticker_data.iloc[i]
                previous_row = ticker_data.iloc[i-1]
                
                current_date = current_row['Date']
                current_multiple = current_row['multiple']
                previous_multiple = previous_row['multiple']
                
                # Check each multiple threshold
                for multiple in self.multiples:
                    # Crossing condition: previous < threshold, current >= threshold
                    if previous_multiple < multiple <= current_multiple:
                        
                        # Check cooldown
                        last_event = last_event_dates[multiple]
                        if last_event is not None:
                            days_since_last = (current_date - last_event).days
                            if days_since_last < self.cooldown_days:
                                continue  # Skip due to cooldown
                        
                        # Record event
                        events.append({
                            'ticker': ticker,
                            'event_date': current_date,
                            'multiple': multiple,
                            'entry_price': current_row['rolling_min_5pct'],
                            'event_price': current_row['Close'],
                            'actual_multiple': current_multiple,
                            'in_sp500': current_row['in_sp500']
                        })
                        
                        # Update last event date
                        last_event_dates[multiple] = current_date
        
        self.events = pd.DataFrame(events)
        
        if len(self.events) > 0:
            print(f"\nDetected {len(self.events):,} events across {self.events['ticker'].nunique()} tickers")
            print("\nEvents by multiple:")
            print(self.events['multiple'].value_counts().sort_index())
        else:
            print("\nWARNING: No events detected!")
            
        return self.events
    
    def calculate_forward_returns(self):
        """
        Calculate forward returns for each event over different time horizons.
        """
        print("\nCalculating forward returns for each event...")
        
        results = []
        
        for idx, event in tqdm(self.events.iterrows(), total=len(self.events), desc="Calculating returns"):
            ticker = event['ticker']
            event_date = event['event_date']
            event_price = event['event_price']
            
            # Get future price data for this ticker
            future_data = self.data[
                (self.data['Ticker'] == ticker) & 
                (self.data['Date'] > event_date)
            ].sort_values('Date').reset_index(drop=True)
            
            if len(future_data) == 0:
                continue
            
            result = {
                'ticker': ticker,
                'event_date': event_date,
                'multiple': event['multiple'],
                'event_price': event_price,
                'entry_price': event['entry_price'],
                'in_sp500': event['in_sp500']
            }
            
            # Calculate metrics for each follow-up period
            for period_name, period_days in self.followup_periods.items():
                # Get data for this period
                period_data = future_data[future_data.index < period_days] if len(future_data) > 0 else pd.DataFrame()
                
                if len(period_data) == 0:
                    # Not enough future data
                    result[f'{period_name}_return'] = np.nan
                    result[f'{period_name}_final_price'] = np.nan
                    result[f'{period_name}_max_price'] = np.nan
                    result[f'{period_name}_min_price'] = np.nan
                    result[f'{period_name}_mdd'] = np.nan
                    result[f'{period_name}_volatility'] = np.nan
                    result[f'{period_name}_days_available'] = 0
                    continue
                
                # Calculate returns
                final_price = period_data.iloc[-1]['Close']
                max_price = period_data['Close'].max()
                min_price = period_data['Close'].min()
                
                total_return = (final_price - event_price) / event_price
                
                # Calculate Maximum Drawdown
                cumulative_returns = (period_data['Close'] / event_price) - 1
                running_max = cumulative_returns.cummax()
                drawdown = cumulative_returns - running_max
                max_drawdown = drawdown.min()
                
                # Calculate volatility (annualized)
                returns = period_data['Close'].pct_change().dropna()
                volatility = returns.std() * np.sqrt(252)  # Annualized
                
                # Store results
                result[f'{period_name}_return'] = total_return
                result[f'{period_name}_final_price'] = final_price
                result[f'{period_name}_max_price'] = max_price
                result[f'{period_name}_min_price'] = min_price
                result[f'{period_name}_mdd'] = max_drawdown
                result[f'{period_name}_volatility'] = volatility
                result[f'{period_name}_days_available'] = len(period_data)
            
            results.append(result)
        
        self.results = pd.DataFrame(results)
        print(f"\nCalculated forward returns for {len(self.results):,} events")
        
        return self.results
    
    def calculate_next_multiple_probability(self):
        """
        Calculate probability of reaching the next multiple.
        Also tracks MDD until next multiple is reached and final returns if not reached.
        """
        print("\nCalculating probability of reaching next multiple...")
        
        next_multiple_results = []
        
        for idx, event in tqdm(self.events.iterrows(), total=len(self.events), desc="Calculating probabilities"):
            ticker = event['ticker']
            event_date = event['event_date']
            current_multiple = event['multiple']
            entry_price = event['entry_price']
            event_price = event['event_price']
            
            # Get future price data
            future_data = self.data[
                (self.data['Ticker'] == ticker) & 
                (self.data['Date'] > event_date)
            ].sort_values('Date').reset_index(drop=True)
            
            if len(future_data) == 0:
                continue
            
            # Check each follow-up period
            for period_name, period_days in self.followup_periods.items():
                period_data = future_data[future_data.index < period_days] if len(future_data) > 0 else pd.DataFrame()
                
                if len(period_data) == 0:
                    continue
                
                # Check if next multiple was reached
                for next_mult in range(current_multiple + 1, 11):
                    target_price = entry_price * next_mult
                    reached = (period_data['Close'] >= target_price).any()
                    
                    if reached:
                        # Find when it was reached
                        first_reach_idx = period_data[period_data['Close'] >= target_price].index[0]
                        days_to_reach = first_reach_idx
                        
                        # Calculate MDD from event to when target was reached
                        data_until_reach = period_data.iloc[:first_reach_idx + 1]
                        cumulative_returns = (data_until_reach['Close'] / event_price) - 1
                        running_max = cumulative_returns.cummax()
                        drawdown = cumulative_returns - running_max
                        mdd_until_reach = drawdown.min()
                        
                        final_return_if_not_reached = np.nan
                    else:
                        days_to_reach = np.nan
                        mdd_until_reach = np.nan
                        
                        # Calculate final return if target was NOT reached
                        final_price = period_data.iloc[-1]['Close']
                        final_return_if_not_reached = (final_price - event_price) / event_price
                    
                    next_multiple_results.append({
                        'ticker': ticker,
                        'event_date': event_date,
                        'current_multiple': current_multiple,
                        'next_multiple': next_mult,
                        'period': period_name,
                        'reached': reached,
                        'days_to_reach': days_to_reach,
                        'mdd_until_reach': mdd_until_reach,
                        'final_return_if_not_reached': final_return_if_not_reached
                    })
        
        self.next_multiple_results = pd.DataFrame(next_multiple_results)
        print(f"Calculated {len(self.next_multiple_results):,} next-multiple checks")
        
        return self.next_multiple_results
    
    def create_summary_statistics(self):
        """
        Create summary statistics table with all KPIs.
        """
        print("\nCreating summary statistics...")
        
        summary_rows = []
        
        for multiple in self.multiples:
            multiple_events = self.results[self.results['multiple'] == multiple]
            
            if len(multiple_events) == 0:
                continue
            
            row = {
                'multiple': f"{multiple}x",
                'n_events': len(multiple_events)
            }
            
            # Calculate statistics for each period
            for period_name in self.followup_periods.keys():
                returns = multiple_events[f'{period_name}_return'].dropna()
                mdds = multiple_events[f'{period_name}_mdd'].dropna()
                
                if len(returns) > 0:
                    # Return percentiles
                    row[f'{period_name}_return_25pct'] = returns.quantile(0.25)
                    row[f'{period_name}_return_50pct'] = returns.quantile(0.50)
                    row[f'{period_name}_return_75pct'] = returns.quantile(0.75)
                    row[f'{period_name}_return_mean'] = returns.mean()
                    
                    # MDD statistics
                    row[f'{period_name}_mdd_mean'] = mdds.mean()
                    row[f'{period_name}_mdd_median'] = mdds.median()
                    row[f'{period_name}_mdd_worst'] = mdds.min()
                    
                    # Win rate (positive returns)
                    row[f'{period_name}_win_rate'] = (returns > 0).mean()
                    
                    # Sharpe ratio (simplified: mean return / volatility)
                    vols = multiple_events[f'{period_name}_volatility'].dropna()
                    if len(vols) > 0:
                        sharpe = (returns.mean() / vols.mean()) if vols.mean() > 0 else np.nan
                        row[f'{period_name}_sharpe'] = sharpe
                    
                    # Sample size
                    row[f'{period_name}_n_samples'] = len(returns)
                else:
                    # Not enough data
                    row[f'{period_name}_return_25pct'] = np.nan
                    row[f'{period_name}_return_50pct'] = np.nan
                    row[f'{period_name}_return_75pct'] = np.nan
                    row[f'{period_name}_return_mean'] = np.nan
                    row[f'{period_name}_mdd_mean'] = np.nan
                    row[f'{period_name}_mdd_median'] = np.nan
                    row[f'{period_name}_mdd_worst'] = np.nan
                    row[f'{period_name}_win_rate'] = np.nan
                    row[f'{period_name}_sharpe'] = np.nan
                    row[f'{period_name}_n_samples'] = 0
                
                # Probability of reaching next multiple
                if hasattr(self, 'next_multiple_results'):
                    next_mult_data = self.next_multiple_results[
                        (self.next_multiple_results['current_multiple'] == multiple) &
                        (self.next_multiple_results['next_multiple'] == multiple + 1) &
                        (self.next_multiple_results['period'] == period_name)
                    ]
                    
                    if len(next_mult_data) > 0:
                        prob = next_mult_data['reached'].mean()
                        row[f'{period_name}_prob_next_multiple'] = prob
                        
                        # Average days to reach (for those that reached)
                        reached_data = next_mult_data[next_mult_data['reached'] == True]
                        if len(reached_data) > 0:
                            avg_days = reached_data['days_to_reach'].mean()
                            row[f'{period_name}_avg_days_to_next'] = avg_days
                            
                            # NEW: MDD until next multiple is reached (percentiles)
                            mdd_until_reach = reached_data['mdd_until_reach'].dropna()
                            if len(mdd_until_reach) > 0:
                                row[f'{period_name}_mdd_until_reach_25pct'] = mdd_until_reach.quantile(0.25)
                                row[f'{period_name}_mdd_until_reach_50pct'] = mdd_until_reach.quantile(0.50)
                                row[f'{period_name}_mdd_until_reach_75pct'] = mdd_until_reach.quantile(0.75)
                        
                        # NEW: Returns if next multiple is NOT reached (percentiles)
                        not_reached_data = next_mult_data[next_mult_data['reached'] == False]
                        if len(not_reached_data) > 0:
                            returns_not_reached = not_reached_data['final_return_if_not_reached'].dropna()
                            if len(returns_not_reached) > 0:
                                row[f'{period_name}_return_not_reached_25pct'] = returns_not_reached.quantile(0.25)
                                row[f'{period_name}_return_not_reached_50pct'] = returns_not_reached.quantile(0.50)
                                row[f'{period_name}_return_not_reached_75pct'] = returns_not_reached.quantile(0.75)
            
            summary_rows.append(row)
        
        self.summary = pd.DataFrame(summary_rows)
        print(f"Created summary table with {len(self.summary)} rows")
        
        return self.summary
    
    def _create_summary_html(self, output_path):
        """
        Create an HTML table from summary statistics with nice formatting.
        
        Parameters:
        -----------
        output_path : Path
            Directory to save the HTML file
        """
        if not hasattr(self, 'summary') or self.summary is None:
            return
        
        # Prepare data for better readability
        df = self.summary.copy()
        
        # Format percentage columns
        pct_cols = [col for col in df.columns if 'return' in col or 'win_rate' in col or 'prob' in col or 'mdd' in col]
        for col in pct_cols:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: f'{x*100:.1f}%' if pd.notna(x) else 'N/A')
        
        # Format days columns
        days_cols = [col for col in df.columns if 'days' in col]
        for col in days_cols:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: f'{x:.0f}' if pd.notna(x) else 'N/A')
        
        # Format sharpe columns
        sharpe_cols = [col for col in df.columns if 'sharpe' in col]
        for col in sharpe_cols:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: f'{x:.2f}' if pd.notna(x) else 'N/A')
        
        # Group columns by period for better organization
        column_groups = {
            'Basic': ['multiple', 'n_events'],
            '1Y': [col for col in df.columns if col.startswith('1Y_')],
            '2Y': [col for col in df.columns if col.startswith('2Y_')],
            '3Y': [col for col in df.columns if col.startswith('3Y_')],
            '5Y': [col for col in df.columns if col.startswith('5Y_')],
            '10Y': [col for col in df.columns if col.startswith('10Y_')]
        }
        
        # Create HTML
        html = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Winner Stocks Analysis - Summary Statistics</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }
        
        h1 {
            color: #2c3e50;
            text-align: center;
            margin-bottom: 10px;
        }
        
        .subtitle {
            text-align: center;
            color: #7f8c8d;
            margin-bottom: 30px;
        }
        
        .controls {
            text-align: center;
            margin-bottom: 20px;
        }
        
        .controls button {
            margin: 0 5px;
            padding: 10px 20px;
            background-color: #3498db;
            color: white;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-size: 14px;
        }
        
        .controls button:hover {
            background-color: #2980b9;
        }
        
        .controls button.active {
            background-color: #2c3e50;
        }
        
        .table-container {
            overflow-x: auto;
            background-color: white;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            padding: 20px;
        }
        
        table {
            border-collapse: collapse;
            width: 100%;
            margin-top: 10px;
        }
        
        th {
            background-color: #34495e;
            color: white;
            padding: 12px 8px;
            text-align: left;
            font-weight: 600;
            position: sticky;
            top: 0;
            z-index: 10;
            border: 1px solid #2c3e50;
        }
        
        td {
            padding: 10px 8px;
            border: 1px solid #ecf0f1;
        }
        
        tr:nth-child(even) {
            background-color: #f8f9fa;
        }
        
        tr:hover {
            background-color: #e8f4f8;
        }
        
        .multiple-col {
            font-weight: bold;
            color: #2c3e50;
            background-color: #ecf0f1 !important;
        }
        
        .period-section {
            display: none;
        }
        
        .period-section.active {
            display: table-cell;
        }
        
        .metric-header {
            font-size: 11px;
            word-wrap: break-word;
        }
        
        .positive {
            color: #27ae60;
        }
        
        .negative {
            color: #e74c3c;
        }
        
        .info-box {
            background-color: #e8f4f8;
            border-left: 4px solid #3498db;
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 5px;
        }
        
        .info-box h3 {
            margin-top: 0;
            color: #2c3e50;
        }
        
        .info-box ul {
            margin: 10px 0;
            padding-left: 20px;
        }
    </style>
</head>
<body>
    <h1>📊 Winner Stocks Analysis</h1>
    <p class="subtitle">Performance after achieving 2x-10x returns</p>
    
    <div class="info-box">
        <h3>Key Metrics Explained</h3>
        <ul>
            <li><strong>Return percentiles (25%, 50%, 75%):</strong> Distribution of returns after reaching each multiple</li>
            <li><strong>Win Rate:</strong> Percentage of cases with positive returns</li>
            <li><strong>MDD (Max Drawdown):</strong> Worst peak-to-trough decline</li>
            <li><strong>Prob Next Multiple:</strong> Probability of reaching the next multiple (e.g., 3x→4x)</li>
            <li><strong>Return if Not Reached:</strong> Returns for cases where next multiple was NOT achieved</li>
            <li><strong>MDD Until Reach:</strong> Drawdown experienced before reaching next multiple</li>
        </ul>
    </div>
    
    <div class="controls">
        <button onclick="showPeriod('all')" class="active" id="btn-all">Show All Periods</button>
        <button onclick="showPeriod('1Y')" id="btn-1Y">1 Year</button>
        <button onclick="showPeriod('2Y')" id="btn-2Y">2 Years</button>
        <button onclick="showPeriod('3Y')" id="btn-3Y">3 Years</button>
        <button onclick="showPeriod('5Y')" id="btn-5Y">5 Years</button>
        <button onclick="showPeriod('10Y')" id="btn-10Y">10 Years</button>
    </div>
    
    <div class="table-container">
        <table>
            <thead>
                <tr>
                    <th rowspan="2" class="multiple-col">Multiple</th>
                    <th rowspan="2">Events</th>
"""
        
        # Add period headers
        for period in ['1Y', '2Y', '3Y', '5Y', '10Y']:
            cols = column_groups[period]
            html += f'                    <th colspan="{len(cols)}" class="period-section {period}" id="header-{period}">{period}</th>\n'
        
        html += """                </tr>
                <tr>
"""
        
        # Add metric headers
        for period in ['1Y', '2Y', '3Y', '5Y', '10Y']:
            for col in column_groups[period]:
                metric_name = col.replace(f'{period}_', '').replace('_', ' ').title()
                html += f'                    <th class="metric-header period-section {period}">{metric_name}</th>\n'
        
        html += """                </tr>
            </thead>
            <tbody>
"""
        
        # Add data rows
        for _, row in df.iterrows():
            html += '                <tr>\n'
            html += f'                    <td class="multiple-col">{row["multiple"]}</td>\n'
            html += f'                    <td>{row["n_events"]}</td>\n'
            
            for period in ['1Y', '2Y', '3Y', '5Y', '10Y']:
                for col in column_groups[period]:
                    value = row.get(col, 'N/A')
                    html += f'                    <td class="period-section {period}">{value}</td>\n'
            
            html += '                </tr>\n'
        
        html += """            </tbody>
        </table>
    </div>
    
    <script>
        function showPeriod(period) {
            // Remove active class from all buttons
            document.querySelectorAll('.controls button').forEach(btn => {
                btn.classList.remove('active');
            });
            
            // Add active class to clicked button
            document.getElementById('btn-' + period).classList.add('active');
            
            if (period === 'all') {
                // Show all periods
                document.querySelectorAll('.period-section').forEach(el => {
                    el.classList.add('active');
                    el.style.display = 'table-cell';
                });
            } else {
                // Hide all periods first
                document.querySelectorAll('.period-section').forEach(el => {
                    el.classList.remove('active');
                    el.style.display = 'none';
                });
                
                // Show selected period
                document.querySelectorAll('.period-section.' + period).forEach(el => {
                    el.classList.add('active');
                    el.style.display = 'table-cell';
                });
            }
        }
        
        // Initialize with all periods visible
        showPeriod('all');
    </script>
</body>
</html>
"""
        
        # Save HTML file
        with open(output_path / 'summary_statistics.html', 'w', encoding='utf-8') as f:
            f.write(html)
    
    def export_ticker_to_excel(self, ticker, output_dir='analysis_results'):
        """
        Export all analysis data for a single ticker to Excel for manual verification.
        
        Parameters:
        -----------
        ticker : str
            The ticker symbol to export
        output_dir : str
            Directory to save the Excel file
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        print(f"\nExporting data for {ticker} to Excel...")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Price History with rolling calculations
        ws1 = wb.create_sheet("Price History")
        ticker_data = self.data[self.data['Ticker'] == ticker].copy()
        ticker_data = ticker_data.sort_values('Date').reset_index(drop=True)
        
        # Calculate multiple for visualization
        if 'rolling_min_5pct' in ticker_data.columns:
            ticker_data['multiple'] = ticker_data['Close'] / ticker_data['rolling_min_5pct']
        
        # Add headers
        headers = ['Date', 'Close', 'Rolling_5pct', 'Multiple', 'In_SP500']
        ws1.append(headers)
        
        # Style header
        for cell in ws1[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for _, row in ticker_data.iterrows():
            ws1.append([
                row['Date'].strftime('%Y-%m-%d') if pd.notna(row['Date']) else '',
                row['Close'] if pd.notna(row['Close']) else '',
                row.get('rolling_min_5pct', '') if pd.notna(row.get('rolling_min_5pct')) else '',
                row.get('multiple', '') if pd.notna(row.get('multiple')) else '',
                row.get('in_sp500', '')
            ])
        
        # Format columns
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 12
        
        # Sheet 2: Detected Events
        ws2 = wb.create_sheet("Events")
        if hasattr(self, 'events'):
            ticker_events = self.events[self.events['ticker'] == ticker].copy()
            
            headers = ['Event_Date', 'Multiple', 'Entry_Price', 'Event_Price', 'Actual_Multiple', 'In_SP500']
            ws2.append(headers)
            
            for cell in ws2[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            for _, row in ticker_events.iterrows():
                ws2.append([
                    row['event_date'].strftime('%Y-%m-%d'),
                    row['multiple'],
                    row['entry_price'],
                    row['event_price'],
                    row['actual_multiple'],
                    row['in_sp500']
                ])
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws2.column_dimensions[col].width = 15
        
        # Sheet 3: Forward Returns by Event
        ws3 = wb.create_sheet("Forward Returns")
        if hasattr(self, 'results'):
            ticker_results = self.results[self.results['ticker'] == ticker].copy()
            
            if len(ticker_results) > 0:
                # Write headers
                headers = ['Event_Date', 'Multiple'] + [
                    f'{period}_{metric}' 
                    for period in ['1Y', '2Y', '3Y', '5Y', '10Y']
                    for metric in ['return', 'final_price', 'max_price', 'mdd']
                ]
                ws3.append(headers)
                
                for cell in ws3[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data
                for _, row in ticker_results.iterrows():
                    data_row = [
                        row['event_date'].strftime('%Y-%m-%d'),
                        row['multiple']
                    ]
                    for period in ['1Y', '2Y', '3Y', '5Y', '10Y']:
                        data_row.extend([
                            row.get(f'{period}_return', ''),
                            row.get(f'{period}_final_price', ''),
                            row.get(f'{period}_max_price', ''),
                            row.get(f'{period}_mdd', '')
                        ])
                    ws3.append(data_row)
                
                # Auto-width for columns
                for col_idx in range(1, len(headers) + 1):
                    ws3.column_dimensions[chr(64 + col_idx)].width = 14
        
        # Sheet 4: Next Multiple Probabilities
        ws4 = wb.create_sheet("Next Multiple Analysis")
        if hasattr(self, 'next_multiple_results'):
            ticker_next = self.next_multiple_results[
                self.next_multiple_results['ticker'] == ticker
            ].copy()
            
            if len(ticker_next) > 0:
                headers = ['Event_Date', 'Current_Multiple', 'Next_Multiple', 'Period', 
                          'Reached', 'Days_to_Reach', 'MDD_Until_Reach', 'Return_If_Not_Reached']
                ws4.append(headers)
                
                for cell in ws4[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                for _, row in ticker_next.iterrows():
                    ws4.append([
                        row['event_date'].strftime('%Y-%m-%d'),
                        row['current_multiple'],
                        row['next_multiple'],
                        row['period'],
                        row['reached'],
                        row.get('days_to_reach', ''),
                        row.get('mdd_until_reach', ''),
                        row.get('final_return_if_not_reached', '')
                    ])
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws4.column_dimensions[col].width = 18
        
        # Save file
        filename = output_path / f'ticker_analysis_{ticker}.xlsx'
        wb.save(filename)
        
        print(f"✓ Exported {ticker} analysis to {filename}")
        
        return filename
    
    def save_results(self, output_dir='analysis_results'):
        """
        Save all results to files.
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        print(f"\nSaving results to {output_dir}/...")
        
        # Save detailed results
        if hasattr(self, 'results'):
            self.results.to_csv(output_path / 'detailed_results.csv', index=False)
            print("âœ“ Saved detailed_results.csv")
        
        # Save summary statistics
        if hasattr(self, 'summary'):
            self.summary.to_csv(output_path / 'summary_statistics.csv', index=False)
            print("âœ“ Saved summary_statistics.csv")
            
            # Also save as HTML
            self._create_summary_html(output_path)
            print("✓ Saved summary_statistics.html")
        
        # Save events
        if hasattr(self, 'events'):
            self.events.to_csv(output_path / 'detected_events.csv', index=False)
            print("âœ“ Saved detected_events.csv")
        
        # Save next multiple probabilities
        if hasattr(self, 'next_multiple_results'):
            self.next_multiple_results.to_csv(output_path / 'next_multiple_probabilities.csv', index=False)
            print("âœ“ Saved next_multiple_probabilities.csv")
        
        print(f"\nAll results saved to {output_dir}/")
    
    def create_visualizations(self, output_dir='analysis_results'):
        """
        Create visualization charts.
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        print("\nCreating visualizations...")
        
        # 1. Return distributions by multiple and period
        self._plot_return_distributions(output_path)
        
        # 2. Probability of reaching next multiple
        self._plot_next_multiple_probability(output_path)
        
        # 3. Maximum Drawdown analysis
        self._plot_drawdown_analysis(output_path)
        
        # 4. Win rate heatmap
        self._plot_win_rate_heatmap(output_path)
        
        # 5. Risk-Return tradeoff
        self._plot_risk_return(output_path)
        
        print("âœ“ All visualizations created")
    
    def _plot_return_distributions(self, output_path):
        """Plot return distributions."""
        fig, axes = plt.subplots(2, 3, figsize=(18, 12))
        fig.suptitle('Return Distributions After Reaching Multiples', fontsize=16, fontweight='bold')
        
        periods = ['1Y', '2Y', '3Y', '5Y', '10Y']
        
        for idx, period in enumerate(periods):
            ax = axes[idx // 3, idx % 3]
            
            data_to_plot = []
            labels = []
            
            for multiple in self.multiples:
                multiple_data = self.results[self.results['multiple'] == multiple]
                returns = multiple_data[f'{period}_return'].dropna()
                
                if len(returns) >= 5:  # Only plot if we have enough data
                    data_to_plot.append(returns.values)
                    labels.append(f'{multiple}x')
            
            if data_to_plot:
                bp = ax.boxplot(data_to_plot, labels=labels, patch_artist=True)
                
                # Color boxes
                colors = plt.cm.RdYlGn(np.linspace(0.3, 0.9, len(data_to_plot)))
                for patch, color in zip(bp['boxes'], colors):
                    patch.set_facecolor(color)
                
                ax.set_title(f'{period} Forward Returns', fontweight='bold')
                ax.set_xlabel('Multiple Achieved')
                ax.set_ylabel('Return (%)')
                ax.axhline(y=0, color='black', linestyle='--', alpha=0.3)
                ax.grid(True, alpha=0.3)
                
                # Format y-axis as percentage
                ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f'{y*100:.0f}%'))
        
        # Remove empty subplot
        fig.delaxes(axes[1, 2])
        
        plt.tight_layout()
        plt.savefig(output_path / 'return_distributions.png', dpi=300, bbox_inches='tight')
        plt.close()
        print("  âœ“ Created return_distributions.png")
    
    def _plot_next_multiple_probability(self, output_path):
        """Plot probability of reaching next multiple."""
        if not hasattr(self, 'next_multiple_results'):
            return
        
        fig, axes = plt.subplots(2, 3, figsize=(18, 12))
        fig.suptitle('Probability of Reaching Next Multiple', fontsize=16, fontweight='bold')
        
        periods = ['1Y', '2Y', '3Y', '5Y', '10Y']
        
        for idx, period in enumerate(periods):
            ax = axes[idx // 3, idx % 3]
            
            probs = []
            labels = []
            
            for multiple in self.multiples[:-1]:  # Exclude 10x as there's no next
                data = self.next_multiple_results[
                    (self.next_multiple_results['current_multiple'] == multiple) &
                    (self.next_multiple_results['next_multiple'] == multiple + 1) &
                    (self.next_multiple_results['period'] == period)
                ]
                
                if len(data) > 0:
                    prob = data['reached'].mean()
                    probs.append(prob * 100)
                    labels.append(f'{multiple}xâ†’{multiple+1}x')
            
            if probs:
                bars = ax.bar(range(len(probs)), probs, color=plt.cm.viridis(np.linspace(0.3, 0.9, len(probs))))
                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha='right')
                ax.set_title(f'{period} Period', fontweight='bold')
                ax.set_ylabel('Probability (%)')
                ax.set_ylim(0, 100)
                ax.grid(True, alpha=0.3, axis='y')
                
                # Add value labels on bars
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           f'{height:.1f}%',
                           ha='center', va='bottom', fontsize=9)
        
        # Remove empty subplot
        fig.delaxes(axes[1, 2])
        
        plt.tight_layout()
        plt.savefig(output_path / 'next_multiple_probability.png', dpi=300, bbox_inches='tight')
        plt.close()
        print("  âœ“ Created next_multiple_probability.png")
    
    def _plot_drawdown_analysis(self, output_path):
        """Plot maximum drawdown analysis."""
        fig, axes = plt.subplots(2, 3, figsize=(18, 12))
        fig.suptitle('Maximum Drawdown Analysis', fontsize=16, fontweight='bold')
        
        periods = ['1Y', '2Y', '3Y', '5Y', '10Y']
        
        for idx, period in enumerate(periods):
            ax = axes[idx // 3, idx % 3]
            
            data_to_plot = []
            labels = []
            
            for multiple in self.multiples:
                multiple_data = self.results[self.results['multiple'] == multiple]
                mdds = multiple_data[f'{period}_mdd'].dropna()
                
                if len(mdds) >= 5:
                    data_to_plot.append(mdds.values * 100)  # Convert to percentage
                    labels.append(f'{multiple}x')
            
            if data_to_plot:
                bp = ax.boxplot(data_to_plot, labels=labels, patch_artist=True)
                
                # Color boxes (red gradient for drawdowns)
                colors = plt.cm.Reds(np.linspace(0.3, 0.9, len(data_to_plot)))
                for patch, color in zip(bp['boxes'], colors):
                    patch.set_facecolor(color)
                
                ax.set_title(f'{period} Maximum Drawdown', fontweight='bold')
                ax.set_xlabel('Multiple Achieved')
                ax.set_ylabel('Maximum Drawdown (%)')
                ax.axhline(y=-50, color='red', linestyle='--', alpha=0.3, label='-50%')
                ax.grid(True, alpha=0.3)
        
        # Remove empty subplot
        fig.delaxes(axes[1, 2])
        
        plt.tight_layout()
        plt.savefig(output_path / 'drawdown_analysis.png', dpi=300, bbox_inches='tight')
        plt.close()
        print("  âœ“ Created drawdown_analysis.png")
    
    def _plot_win_rate_heatmap(self, output_path):
        """Plot win rate heatmap."""
        # Prepare data for heatmap
        win_rates = []
        
        for multiple in self.multiples:
            row = []
            for period in ['1Y', '2Y', '3Y', '5Y', '10Y']:
                multiple_data = self.results[self.results['multiple'] == multiple]
                returns = multiple_data[f'{period}_return'].dropna()
                
                if len(returns) > 0:
                    win_rate = (returns > 0).mean() * 100
                    row.append(win_rate)
                else:
                    row.append(np.nan)
            
            win_rates.append(row)
        
        win_rates_df = pd.DataFrame(
            win_rates,
            index=[f'{m}x' for m in self.multiples],
            columns=['1Y', '2Y', '3Y', '5Y', '10Y']
        )
        
        plt.figure(figsize=(10, 8))
        sns.heatmap(win_rates_df, annot=True, fmt='.1f', cmap='RdYlGn', 
                   vmin=0, vmax=100, cbar_kws={'label': 'Win Rate (%)'})
        plt.title('Win Rate (% Positive Returns) After Reaching Multiple', 
                 fontsize=14, fontweight='bold', pad=20)
        plt.xlabel('Time Period', fontweight='bold')
        plt.ylabel('Multiple Achieved', fontweight='bold')
        plt.tight_layout()
        plt.savefig(output_path / 'win_rate_heatmap.png', dpi=300, bbox_inches='tight')
        plt.close()
        print("  âœ“ Created win_rate_heatmap.png")
    
    def _plot_risk_return(self, output_path):
        """Plot risk-return scatter for different holding periods."""
        fig, axes = plt.subplots(2, 3, figsize=(18, 12))
        fig.suptitle('Risk-Return Tradeoff by Multiple and Period', fontsize=16, fontweight='bold')
        
        periods = ['1Y', '2Y', '3Y', '5Y', '10Y']
        
        for idx, period in enumerate(periods):
            ax = axes[idx // 3, idx % 3]
            
            for multiple in self.multiples:
                multiple_data = self.results[self.results['multiple'] == multiple]
                returns = multiple_data[f'{period}_return'].dropna()
                mdds = multiple_data[f'{period}_mdd'].dropna()
                
                # Get matching indices
                valid_indices = returns.index.intersection(mdds.index)
                
                if len(valid_indices) >= 3:
                    avg_return = returns.loc[valid_indices].mean() * 100
                    avg_mdd = abs(mdds.loc[valid_indices].mean()) * 100
                    
                    ax.scatter(avg_mdd, avg_return, s=200, alpha=0.6, 
                             label=f'{multiple}x', edgecolors='black', linewidth=1)
                    ax.text(avg_mdd, avg_return, f'{multiple}x', 
                           ha='center', va='center', fontsize=9, fontweight='bold')
            
            ax.set_title(f'{period} Period', fontweight='bold')
            ax.set_xlabel('Avg Max Drawdown (%)', fontweight='bold')
            ax.set_ylabel('Avg Return (%)', fontweight='bold')
            ax.axhline(y=0, color='black', linestyle='--', alpha=0.3)
            ax.grid(True, alpha=0.3)
        
        # Remove empty subplot
        fig.delaxes(axes[1, 2])
        
        plt.tight_layout()
        plt.savefig(output_path / 'risk_return_scatter.png', dpi=300, bbox_inches='tight')
        plt.close()
        print("  âœ“ Created risk_return_scatter.png")
    
    def run_full_analysis(self, output_dir='analysis_results'):
        """
        Run the complete analysis pipeline.
        """
        print("\n" + "="*70)
        print("WINNER STOCKS ANALYSIS")
        print("="*70)
        
        # Step 1: Calculate rolling entry prices
        self.calculate_rolling_entry_price()
        
        # Step 2: Detect multiple events
        self.detect_multiple_events()
        
        if len(self.events) == 0:
            print("\nERROR: No events detected. Cannot continue analysis.")
            return
        
        # Step 3: Calculate forward returns
        self.calculate_forward_returns()
        
        # Step 4: Calculate next multiple probabilities
        self.calculate_next_multiple_probability()
        
        # Step 5: Create summary statistics
        self.create_summary_statistics()
        
        # Step 6: Save results
        self.save_results(output_dir)
        
        # Step 7: Create visualizations
        self.create_visualizations(output_dir)
        
        print("\n" + "="*70)
        print("ANALYSIS COMPLETE!")
        print("="*70)
        print(f"\nResults saved to: {output_dir}/")
        print("\nGenerated files:")
        print("  - summary_statistics.csv (main results table)")
        print("  - summary_statistics.html (interactive HTML table)")
        print("  - detailed_results.csv (all individual events)")
        print("  - detected_events.csv (all crossing events)")
        print("  - next_multiple_probabilities.csv")
        print("  - return_distributions.png")
        print("  - next_multiple_probability.png")
        print("  - drawdown_analysis.png")
        print("  - win_rate_heatmap.png")
        print("  - risk_return_scatter.png")
        print("\nTo export data for a specific ticker to Excel:")
        print("  analyzer.export_ticker_to_excel('AAPL')")
        
        return self.summary


def main():
    """
    Main function to run analysis.
    """
    from download_sp500_data import SP500DataManager
    
    print("Loading S&P 500 historical data...")
    
    # Load data
    manager = SP500DataManager()
    df, metadata = manager.load_data()
    
    if df is None:
        print("ERROR: Could not load data. Please run download_sp500_data.py first.")
        return
    
    # Only analyze stocks that were in S&P 500
    print(f"\nFiltering for S&P 500 members only...")
    df_sp500 = df[df['in_sp500'] == True].copy()
    print(f"Rows after filtering: {len(df_sp500):,} (from {len(df):,})")
    
    # Run analysis
    analyzer = WinnerStocksAnalyzer(df_sp500)
    summary = analyzer.run_full_analysis()
    
    # Display summary
    if summary is not None:
        print("\n" + "="*70)
        print("SUMMARY STATISTICS PREVIEW")
        print("="*70)
        print("\nFirst few rows of summary table:")
        print(summary.head(10).to_string())
        
        # Example: Export a single ticker for manual verification
        # Uncomment the following lines to export a specific ticker:
        # print("\n" + "="*70)
        # print("EXAMPLE: Exporting single ticker data")
        # print("="*70)
        # if len(analyzer.events) > 0:
        #     example_ticker = analyzer.events.iloc[0]['ticker']
        #     analyzer.export_ticker_to_excel(example_ticker)
        #     print(f"\nTo export another ticker, use:")
        #     print(f"  analyzer.export_ticker_to_excel('YOUR_TICKER')")


if __name__ == "__main__":
    main()
